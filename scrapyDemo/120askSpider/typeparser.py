# -*- coding: utf-8 -*-
"""
Created on Sun Nov 17 2020
"""
__author__ = 'jingzl'
__version__ = '1.0'


import os
import pandas as pd
import numpy as np
import xlrd
import time
import datetime
from openpyxl import load_workbook

CONST_DATA_PATH = './data/'
CONST_OUTPUT_PATH = './output/'


def logError(url):
    print('error: ' + url)


def append_df_to_excel(excel_file, dl):
    writer = pd.ExcelWriter(excel_file, engine='openpyxl', options={'strings_to_urls': False})
    writer.book = load_workbook(excel_file)
    start_row = writer.book['Sheet1'].max_row
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    df = pd.DataFrame(dl)
    df.to_excel(writer, 'Sheet1', startrow=start_row, header=False, index=False)
    writer.save()

'''
def outputURL(url1, url2):
    dl = []
    dl.append([url2])
    fname = CONST_OUTPUT_PATH + url1[0] + '_' + url1[1] + '.xlsx'
    if not os.path.exists(fname):
        df = pd.DataFrame(dl, index=np.arange(len(dl)))
        df.to_excel(fname, header=False, index=False)
    else:
        append_df_to_excel(fname, dl)
'''


def outputURL(dl):
    for item in dl:
        ls = list(dl[item])
        url = item.split('//')
        url1 = url[1].split('/')
        fname = CONST_OUTPUT_PATH + url[0].split(':')[0] + '_' + url1[0] + '_' + url1[1] + '.xlsx'
        df = pd.DataFrame(ls)
        with pd.ExcelWriter(fname, engine='xlsxwriter', options={'strings_to_urls': False}) as writer:
            df.to_excel(writer, header=False, index=False)


def parseTypeForURL(typeFile, files):
    try:
        typels = []
        typeExcel = xlrd.open_workbook(CONST_DATA_PATH+typeFile, encoding_override='utf-8')
        if typeExcel:
            sheet = typeExcel.sheet_by_index(0)
            for i in range(sheet.nrows):
                row = sheet.row(i)
                typeurl = row[0].value.strip()
                typels.append(typeurl)

        dl = {}
        for i in range(len(files)):
            datafile = CONST_DATA_PATH + files[i]
            print(datafile)
            excel = xlrd.open_workbook(datafile, encoding_override='utf-8')
            if excel:
                sheet = excel.sheet_by_index(0)
                for j in range(sheet.nrows):
                    row = sheet.row(j)
                    url = row[0].value.strip()
                    try:
                        url = url.split(' ')[0]
                        url1 = url.split('//')
                        url2 = url1[1].split('/')
                        url3 = url1[0] + '//' + url2[0] + '/' + url2[1]
                        if url3 in typels:
                            dl.setdefault(url3, set()).add(url)
                    except Exception as e:
                        print(e)
                        logError(url)
                        continue
        outputURL(dl)

    except Exception as e:
        print(e)


def parseSheet(sheet):
    type_list = []
    for i in range(sheet.nrows):
        row = sheet.row(i)
        url = row[0].value.strip()
        #print(url)
        # https://m.120ask.com/shilu/038cvwlcawkim7p9
        try:
            url1 = url.split('//')
            url2 = url1[1].split('/')
            url3 = url1[0] + '//' + url2[0] + '/' + url2[1]
            type_list.append(url3)
        except Exception as e:
            print(e)
            logError(url)
            continue
    return type_list


def parseFileForType(files):
    try:
        dl = []
        for i in range(len(files)):
            datafile = CONST_DATA_PATH + files[i]
            print(datafile)
            excel = xlrd.open_workbook(datafile, encoding_override='utf-8')
            if excel:
                print(datafile)
                # all_sheet = excel.sheets()
                # print(all_sheet)
                sheet = excel.sheet_by_index(0)
                dl += parseSheet(sheet)
        res = pd.value_counts(dl)
        dt = {'分类': res.index, '个数': res.values}
        df = pd.DataFrame(dt)
        df.to_excel(CONST_OUTPUT_PATH + '分类统计_{0}.xlsx'.format(datetime.datetime.now().strftime('%Y%m%d%H%M%S')),
            index=False)
    except Exception as e:
        print(e)


if __name__ == '__main__':
    print("开始处理--{0}".format(datetime.datetime.now().strftime('%Y%m%d %H:%M:%S')).center(100, '-'))
    start = time.perf_counter()

    # 解析原始数据统计分类 v1-120ask_test.xlsx
    firstFiles = ['v1-120ask.xlsx', 'v2-120ask.xlsx', 'v3-120ask.xlsx']
    #print("[统计分类信息]-{0}".format(datetime.datetime.now().strftime('%Y%m%d %H:%M:%S')).center(100, '-'))
    #parseFileForType(firstFiles)

    # 手工筛选，确认要爬取的分类，并根据分类文件拆分URL
    typeFile = '分类统计_20201119_FINAL.xlsx'
    print("[根据确认分类提取URL]-{0}".format(datetime.datetime.now().strftime('%Y%m%d %H:%M:%S')).center(100, '-'))
    parseTypeForURL(typeFile, firstFiles)

    # 3. 根据URL的统计量倒序排列，优先处理量大的类型
    # 4. 分别将不同类型的URL爬取，并提取数据存储（每一种类型都需要单独分析，爬取）
    # 5. 根据不同类型分别输出新格式的excel
    # 6. 提供结果数据的最终合并操作

    dur = time.perf_counter() - start
    print("总计用时：{:.2f}s".format(dur))
    print("处理完毕--{0}".format(datetime.datetime.now().strftime('%Y%m%d %H:%M:%S')).center(100, '-'))
    # end
    print("all end".center(100, '-'))
